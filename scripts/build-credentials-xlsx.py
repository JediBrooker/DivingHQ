#!/usr/bin/env python3
"""Render docs/seed-credentials.csv into a formatted docs/seed-credentials.xlsx.

The CSV is the canonical artifact (emitted by scripts/generate-seed.js, zero deps).
This optional helper upgrades it to a presentation spreadsheet. It needs openpyxl,
which is NOT a project dependency — install into a throwaway venv:

    python3 -m venv /tmp/xlsxenv && /tmp/xlsxenv/bin/pip install openpyxl
    /tmp/xlsxenv/bin/python scripts/build-credentials-xlsx.py
"""
import csv
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CSV_PATH = os.path.join(ROOT, "docs", "seed-credentials.csv")
XLSX_PATH = os.path.join(ROOT, "docs", "seed-credentials.xlsx")

HEADERS = ["persona", "username", "password", "federation", "home_org", "full_name", "email", "notes"]
TITLES = ["Persona", "Username", "Password", "Federation", "Home Org", "Full Name", "Email", "Notes"]
WIDTHS = [22, 18, 14, 24, 18, 22, 34, 52]

NAVY = "1F3A5F"
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
NOTE_FILL = PatternFill("solid", fgColor="FFF3CD")   # pale amber: rows that need attention
ZEBRA_FILL = PatternFill("solid", fgColor="F2F6FA")
thin = Side(style="thin", color="D7DEE6")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

with open(CSV_PATH, newline="", encoding="utf-8") as f:
    rows = list(csv.DictReader(f))

wb = Workbook()
ws = wb.active
ws.title = "Logins"

# Title banner
ws.merge_cells("A1:H1")
t = ws["A1"]
t.value = "DivingHQ — seed login credentials  (every account uses password: password123, except admin/admin)"
t.font = Font(bold=True, size=13, color="FFFFFF")
t.fill = PatternFill("solid", fgColor=NAVY)
t.alignment = Alignment(vertical="center", horizontal="left", indent=1)
ws.row_dimensions[1].height = 26

# Header row
for c, title in enumerate(TITLES, start=1):
    cell = ws.cell(row=2, column=c, value=title)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    cell.border = BORDER
ws.row_dimensions[2].height = 20

# Data rows
for i, row in enumerate(rows):
    r = i + 3
    has_note = bool(row.get("notes"))
    for c, key in enumerate(HEADERS, start=1):
        cell = ws.cell(row=r, column=c, value=row.get(key, ""))
        cell.alignment = Alignment(vertical="center", horizontal="left", indent=1, wrap_text=(key == "notes"))
        cell.border = BORDER
        if key in ("username", "password"):
            cell.font = Font(name="Menlo", size=11, bold=(key == "password"))
        if has_note:
            cell.fill = NOTE_FILL
        elif i % 2 == 1:
            cell.fill = ZEBRA_FILL

for c, w in enumerate(WIDTHS, start=1):
    ws.column_dimensions[get_column_letter(c)].width = w

ws.freeze_panes = "A3"
ws.auto_filter.ref = f"A2:H{len(rows) + 2}"

# Read-me sheet
info = wb.create_sheet("Read me")
lines = [
    ("DivingHQ seed — how to use these logins", True),
    ("", False),
    ("Password: every seeded account uses  password123 .  The super-admin is  admin / admin .", False),
    ("All emails are pre-verified, so any persona can sign in immediately at /login.", False),
    ("", False),
    ("Two federations:  Diving Australia (AUS)  and  British Aquatic Sports (GBR).", False),
    ("Per federation: 1 org admin, 1 meet manager, 1 referee, 20 divers, 2 coaches, 2 spectators.", False),
    ("11 judges (judge.01–judge.11) are shared across BOTH federations (they cover 11-judge synchro panels).", False),
    ("", False),
    ("Worth a look (highlighted rows):", True),
    ("  • judge.03 — scores erratically; open Judge Analysis to see the outliers surface.", False),
    ("  • aus.diver.20 — suspended; use it to test reactivation in the User Manager.", False),
    ("  • gbr.diver.05 — has a pending club-change request in the org-admin queue.", False),
    ("  • aus.diver.07 — mid cross-federation transfer (source approved, target pending).", False),
    ("  • aus.fan.01 — has a pending role request (judge).", False),
    ("", False),
    ("Data: 5 meets / 25 events over 3 years — mostly Completed (results, recaps, records,", False),
    ("Judge Analysis), plus Live events (scoreboard + control room) and Upcoming events (registration).", False),
]
for i, (text, bold) in enumerate(lines, start=1):
    cell = info.cell(row=i, column=1, value=text)
    cell.font = Font(bold=bold, size=12 if bold and i == 1 else 11)
info.column_dimensions["A"].width = 110

wb.save(XLSX_PATH)
print(f"wrote {os.path.relpath(XLSX_PATH, ROOT)}  ({len(rows)} logins)")
